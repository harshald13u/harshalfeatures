#!/usr/bin/env python3
"""Single source of truth -> site data.
Reads the master Excel (FII_DII_History.xlsx, sheet 'Data') and writes
fii-dii/history.json in the shape the FII/DII page expects.
Optionally merges/back-fills from a CSV (NSE / Trendlyne / any) passed as an arg:
    python3 fii-dii/_tools/build_fii_history.py [path/to/extra.csv]
CSV is matched by header keywords (date / fii net / dii net / buy / sell), upserted by date.
Run after every data change (a daily scheduled task does this automatically)."""
import sys, os, json, csv, datetime, re

ROOT  = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
# master Excel lives in the user's workspace folder (one level into the mounted folder)
XLSX_CANDIDATES = [
    os.path.join(ROOT, "FII_DII_History.xlsx"),
    "/sessions/great-dreamy-darwin/mnt/Features/FII_DII_History.xlsx",
]
OUT = os.path.join(ROOT, "fii-dii", "history.json")

def find_xlsx():
    for p in XLSX_CANDIDATES:
        if os.path.exists(p): return p
    return None

def num(v):
    if v is None or v=="" : return None
    s=str(v).replace(",","").replace("₹","").strip()
    if s in ("-","—",""): return None
    try: return float(s)
    except: return None

def norm_date(v):
    if isinstance(v,(datetime.date,datetime.datetime)): return v.strftime("%Y-%m-%d")
    s=str(v).strip()
    for f in ("%Y-%m-%d","%d-%b-%Y","%d %b %Y","%d-%m-%Y","%d/%m/%Y","%m/%d/%Y","%d-%B-%Y"):
        try: return datetime.datetime.strptime(s,f).strftime("%Y-%m-%d")
        except: pass
    return None

def load_excel(path):
    import openpyxl
    wb=openpyxl.load_workbook(path, data_only=True); ws=wb["Data"] if "Data" in wb.sheetnames else wb.active
    rows=list(ws.iter_rows(values_only=True)); hdr=[str(h).strip().lower() if h else "" for h in rows[0]]
    def col(*names):
        for n in names:
            if n in hdr: return hdr.index(n)
        return None
    ci={"date":col("date"),"fb":col("fii buy"),"fs":col("fii sell"),"fn":col("fii net"),
        "db":col("dii buy"),"ds":col("dii sell"),"dn":col("dii net")}
    data={}
    for r in rows[1:]:
        if not any(r): continue
        d=norm_date(r[ci["date"]]) if ci["date"] is not None else None
        if not d: continue
        g=lambda k: num(r[ci[k]]) if ci[k] is not None and ci[k]<len(r) else None
        data[d]={"fb":g("fb"),"fs":g("fs"),"fn":g("fn"),"db":g("db"),"ds":g("ds"),"dn":g("dn")}
    return data

def merge_csv(data, path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        rd=csv.reader(f); rows=[r for r in rd if any(c.strip() for c in r)]
    if not rows: return 0
    hdr=[c.strip().lower() for c in rows[0]]
    def col(*subs):
        for i,h in enumerate(hdr):
            if all(s in h for s in subs): return i
        return None
    ci={"date":col("date"),"fb":col("fii","buy"),"fs":col("fii","sell"),"fn":col("fii","net"),
        "db":col("dii","buy"),"ds":col("dii","sell"),"dn":col("dii","net")}
    if ci["date"] is None: print("  CSV: no date column found, skipping"); return 0
    n=0
    for r in rows[1:]:
        d=norm_date(r[ci["date"]]) if ci["date"]<len(r) else None
        if not d: continue
        g=lambda k: num(r[ci[k]]) if ci[k] is not None and ci[k]<len(r) else None
        rec={"fb":g("fb"),"fs":g("fs"),"fn":g("fn"),"db":g("db"),"ds":g("ds"),"dn":g("dn")}
        if rec["fn"] is None and rec["fb"] is not None and rec["fs"] is not None: rec["fn"]=rec["fb"]-rec["fs"]
        if rec["dn"] is None and rec["db"] is not None and rec["ds"] is not None: rec["dn"]=rec["db"]-rec["ds"]
        data[d]=rec; n+=1
    return n

def write_excel_back(path, data):
    import openpyxl
    wb=openpyxl.load_workbook(path); ws=wb["Data"] if "Data" in wb.sheetnames else wb.active
    # clear existing data rows
    if ws.max_row>1: ws.delete_rows(2, ws.max_row-1)
    for d in sorted(data):
        r=data[d]
        ws.append([datetime.date.fromisoformat(d), r.get("fb"), r.get("fs"), r.get("fn"),
                   r.get("db"), r.get("ds"), r.get("dn"), "compiled","provisional"])
    for row in range(2, ws.max_row+1): ws.cell(row=row,column=1).number_format="yyyy-mm-dd"
    wb.save(path)

def main():
    xlsx=find_xlsx()
    if not xlsx: raise SystemExit("FII_DII_History.xlsx not found")
    data=load_excel(xlsx)
    if len(sys.argv)>1 and os.path.exists(sys.argv[1]):
        added=merge_csv(data, sys.argv[1]); print(f"  merged {added} rows from {sys.argv[1]}")
        write_excel_back(xlsx, data)
    H=[]
    for d in sorted(data):
        r=data[d]
        fn=r.get("fn"); dn=r.get("dn")
        if fn is None and r.get("fb") is not None and r.get("fs") is not None: fn=r["fb"]-r["fs"]
        if dn is None and r.get("db") is not None and r.get("ds") is not None: dn=r["db"]-r["ds"]
        if fn is None or dn is None: continue
        fii={"net":fn}; dii={"net":dn}
        if r.get("fb") is not None: fii.update(buy=r["fb"],sell=r["fs"])
        if r.get("db") is not None: dii.update(buy=r["db"],sell=r["ds"])
        H.append({"date":d,"fii":fii,"dii":dii})
    last=H[-1]
    out={"lastUpdated":datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
         "latest":{"date":last["date"],"status":"provisional","confidence":"confirmed",
                   "sources":["Master data · NSE-compiled"],
                   "fii":last["fii"] if "buy" in last["fii"] else {"buy":0,"sell":0,"net":last["fii"]["net"]},
                   "dii":last["dii"] if "buy" in last["dii"] else {"buy":0,"sell":0,"net":last["dii"]["net"]}},
         "history":H}
    # include the monthly series so Monthly/Yearly tabs keep 2007-2026
    import csv as _csv
    MCSV=os.path.join(ROOT,"fii-dii","fii_dii_monthly.csv")
    if os.path.exists(MCSV):
        M=[]
        with open(MCSV,newline="",encoding="utf-8-sig") as fh:
            for row in _csv.DictReader(fh):
                ym=(row.get("Month") or "").strip()[:7]
                if len(ym)!=7: continue
                fn=num(row.get("FII Net")); dn=num(row.get("DII Net"))
                if fn is None or dn is None: continue
                fo={"net":fn}; do={"net":dn}
                if num(row.get("FII Buy")) is not None: fo.update(buy=num(row["FII Buy"]),sell=num(row["FII Sell"]))
                if num(row.get("DII Buy")) is not None: do.update(buy=num(row["DII Buy"]),sell=num(row["DII Sell"]))
                M.append({"date":ym+"-01","fii":fo,"dii":do})
        out["monthly"]=M
    json.dump(out, open(OUT,"w"), separators=(",",":"))
    print(f"wrote {OUT}: {len(H)} trading days ({H[0]['date']} to {H[-1]['date']})")

if __name__=="__main__": main()
