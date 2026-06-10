#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Auto-update the FII / FPI Historical Flows tool with the latest month's data.

WHAT IT DOES
  1. Reads the embedded  const D = {...};  block out of tools/fii-flows/index.html
  2. Fetches the latest monthly FPI net-investment figures from NSDL
     (Calendar-year report, Rs-Crore columns: Equity + Total).
  3. Upserts any NEW months (and refreshes the most-recent, still-provisional month)
     into D.monthly.
  4. Re-derives the CURRENT calendar-year row (D.cy) and CURRENT financial-year row
     (D.fy) from the monthly sums, and recomputes the FY cumulative chain.
  5. Writes back:  index.html  (inline D),  fii-flows-data.json (mirror),
     and regenerates  FII_FPI_Historical_Data_India.xlsx  for download.

SAFETY
  * Strong validation gates. If anything looks implausible or the page can't be
    parsed, it raises SystemExit(non-zero) and writes NOTHING — so an unattended
    cron run can NEVER commit garbage; it just fails loudly.
  * Idempotent: re-running with no new month exits 0 and changes nothing.

USAGE
  python3 update_fii_flows.py            # fetch NSDL, update files
  python3 update_fii_flows.py --dry-run  # parse + validate, print, write nothing
  python3 update_fii_flows.py --from-csv path.csv   # update from a CSV instead of NSDL
        CSV headers (case-insensitive, keyword match): month/period, equity, total
        month accepts: 2026-05, May-2026, May 2026, 01-05-2026, etc.

Deps: requests, beautifulsoup4, openpyxl
"""
import os, re, sys, json, time, datetime, calendar

HERE = os.path.dirname(os.path.abspath(__file__))
FLOW = os.path.abspath(os.path.join(HERE, ".."))          # tools/fii-flows
INDEX = os.path.join(FLOW, "index.html")
JSON_MIRROR = os.path.join(FLOW, "fii-flows-data.json")
XLSX_OUT = os.path.join(FLOW, "FII_FPI_Historical_Data_India.xlsx")

CEIL_MONTH = 300000      # |eq| or |tot| for a single month, Rs cr  (sanity ceiling)
CEIL_YEAR  = 2000000     # |tot| for a full year, Rs cr
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")
MONTHS = {m.lower():i for i,m in enumerate(calendar.month_name) if m}
MONTHS.update({m.lower():i for i,m in enumerate(calendar.month_abbr) if m})

DRY = "--dry-run" in sys.argv

# ---------------------------------------------------------------- helpers
def log(*a): print(*a, flush=True)

def num(s):
    """Parse an Indian-formatted number; (x)=negative; '' / - => None."""
    if s is None: return None
    t = str(s).strip().replace("₹","").replace(",","").replace("\xa0","")
    neg = False
    if t.startswith("(") and t.endswith(")"): neg, t = True, t[1:-1]
    t = t.replace("−","-")          # unicode minus
    if t in ("","-","—","NA","N.A.","—"): return None
    try:
        v = float(t)
    except ValueError:
        return None
    return -v if neg else v

def ym_of(text):
    """'May 2026' / 'May-26' / '2026-05' -> '2026-05' (or None)."""
    s = str(text).strip()
    m = re.match(r"^(\d{4})-(\d{1,2})$", s)
    if m: return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    m = re.search(r"([A-Za-z]{3,9})[\s\-/]+(\d{2,4})", s)
    if m and m.group(1).lower() in MONTHS:
        mo = MONTHS[m.group(1).lower()]; yr = int(m.group(2))
        if yr < 100: yr += 2000
        return f"{yr:04d}-{mo:02d}"
    for fmt in ("%d-%m-%Y","%d/%m/%Y","%Y-%m"):
        try: return datetime.datetime.strptime(s, fmt).strftime("%Y-%m")
        except ValueError: pass
    return None

# ---------------------------------------------------------------- read current D
def read_D():
    html = open(INDEX, encoding="utf-8").read()
    m = re.search(r"const D=(\{.*?\});", html, re.S)
    if not m: raise SystemExit("FATAL: could not locate `const D={...};` in index.html")
    return html, m.span(1), json.loads(m.group(1))

# ---------------------------------------------------------------- NSDL fetch
def fetch_nsdl_months():
    """Return {ym: {'eq':float,'tot':float}} for recent months from NSDL CY report."""
    import requests
    from bs4 import BeautifulSoup
    urls = [
        "https://www.fpi.nsdl.co.in/web/Reports/Yearwise.aspx?RptType=6",
        "https://www.fpi.nsdl.co.in/Reports/Yearwise.aspx?RptType=6",
    ]
    sess = requests.Session()
    sess.headers.update({"User-Agent": UA, "Accept-Language": "en-US,en;q=0.9",
                         "Accept": "text/html,application/xhtml+xml"})
    last_err = None
    for url in urls:
        for attempt in range(3):
            try:
                r = sess.get(url, timeout=40); r.raise_for_status()
                out = parse_nsdl(r.text)
                if out: return out
                last_err = "parsed 0 month-rows"
            except Exception as e:
                last_err = repr(e)
            time.sleep(3)
    raise SystemExit(f"FATAL: NSDL fetch/parse failed ({last_err}). "
                     f"Nothing written. (HTML layout may have changed — adjust parse_nsdl.)")

def parse_nsdl(html):
    """Header-mapped parse of the NSDL GridView. Picks Rs-crore Equity + Total columns."""
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "html.parser")
    best = {}
    for tbl in soup.find_all("table"):
        rows = tbl.find_all("tr")
        if len(rows) < 3: continue
        # locate a header row that names Equity & Total
        hdr_idx = eq_col = tot_col = None
        for ri, tr in enumerate(rows[:4]):
            cells = [c.get_text(" ", strip=True).lower() for c in tr.find_all(["th","td"])]
            if not cells: continue
            ec = tc = None
            for ci, h in enumerate(cells):
                if "usd" in h or "us$" in h or "$" in h: continue   # skip dollar columns
                if ec is None and "equity" in h: ec = ci
                if "total" in h or "grand total" in h or h == "net": tc = ci
            if ec is not None and tc is not None:
                hdr_idx, eq_col, tot_col = ri, ec, tc; break
        if hdr_idx is None: continue
        got = {}
        for tr in rows[hdr_idx+1:]:
            cells = [c.get_text(" ", strip=True) for c in tr.find_all(["td","th"])]
            if len(cells) <= max(eq_col, tot_col): continue
            ym = ym_of(cells[0])
            if not ym: continue
            eq = num(cells[eq_col]); tot = num(cells[tot_col])
            if eq is None or tot is None: continue
            got[ym] = {"eq": round(eq), "tot": round(tot)}
        if len(got) > len(best): best = got
    return best

# ---------------------------------------------------------------- CSV fetch
def fetch_csv(path):
    import csv
    out = {}
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = [r for r in csv.reader(f) if any(c.strip() for c in r)]
    if not rows: raise SystemExit("FATAL: empty CSV")
    hdr = [c.strip().lower() for c in rows[0]]
    def col(*subs):
        for i,h in enumerate(hdr):
            if any(s in h for s in subs): return i
        return None
    mi, ei, ti = col("month","period","ym","date"), col("equity","eq"), col("total","tot","net")
    if mi is None or ei is None or ti is None:
        raise SystemExit(f"FATAL: CSV needs month/equity/total columns; got {hdr}")
    for r in rows[1:]:
        ym = ym_of(r[mi]) if mi < len(r) else None
        if not ym: continue
        eq, tot = num(r[ei]), num(r[ti])
        if eq is None or tot is None: continue
        out[ym] = {"eq": round(eq), "tot": round(tot)}
    return out

# ---------------------------------------------------------------- merge + roll-up
def fy_of(ym):
    y, mo = int(ym[:4]), int(ym[5:7])
    start = y if mo >= 4 else y-1
    return f"{start}-{str(start+1)[2:]}"

def apply_updates(D, fresh):
    monthly = D["monthly"]
    have = {r["ym"]: r for r in monthly}
    last_ym = max(have) if have else "0000-00"
    now = datetime.date.today()
    cur_ym = f"{now.year:04d}-{now.month:02d}"
    changed = []
    for ym in sorted(fresh):
        if ym > cur_ym:                      # never accept a future month
            continue
        rec = fresh[ym]
        if abs(rec["eq"]) > CEIL_MONTH or abs(rec["tot"]) > CEIL_MONTH:
            raise SystemExit(f"FATAL: implausible month {ym}: {rec} (>|{CEIL_MONTH}|). Nothing written.")
        # accept if new, OR refresh the most-recent two months (still provisional/revised)
        if ym not in have:
            if ym <= last_ym:                # gap/back-fill we didn't expect -> skip silently
                continue
            monthly.append({"ym": ym, "eq": rec["eq"], "tot": rec["tot"]}); changed.append(ym)
        else:
            old = have[ym]
            if (old.get("eq"), old.get("tot")) != (rec["eq"], rec["tot"]) and ym >= last_ym:
                old["eq"], old["tot"] = rec["eq"], rec["tot"]; changed.append(ym+"*")
    monthly.sort(key=lambda r: r["ym"])
    if not changed:
        return changed
    rebuild_year_rows(D)
    return changed

def rebuild_year_rows(D):
    monthly = D["monthly"]
    # ---- current & touched calendar years ----
    cy_by = {}
    for r in monthly:
        y = int(r["ym"][:4]); d = cy_by.setdefault(y, {"eq":0,"tot":0,"m":0})
        d["eq"] += r["eq"]; d["tot"] += r["tot"]; d["m"] += 1
    cy = {r["y"]: r for r in D["cy"]}
    for y in sorted(cy_by):
        if y not in cy:        # only auto-extend forward; never rewrite deep history
            if y <= max(cy):  # historic gap -> leave the official figure alone
                continue
            cy[y] = {"y": y}; D["cy"].append(cy[y])
        # refresh ONLY the latest two CY rows (current + prior partial->final), from monthly sums
        if y >= max(cy_by)-1:
            agg = cy_by[y]
            cy[y].update(y=y, eq=agg["eq"], dt=agg["tot"]-agg["eq"], tot=agg["tot"], m=agg["m"])
    D["cy"].sort(key=lambda r: r["y"])
    # ---- current & touched financial years ----
    fy_by = {}
    for r in monthly:
        k = fy_of(r["ym"]); d = fy_by.setdefault(k, {"eq":0,"tot":0})
        d["eq"] += r["eq"]; d["tot"] += r["tot"]
    fy = {r["fy"]: r for r in D["fy"]}
    def fy_start(s): return int(s[:4])
    maxfy = max(fy_by, key=fy_start)
    for k in sorted(fy_by, key=fy_start):
        if k not in fy:
            if fy_start(k) <= max(fy_start(x) for x in fy): continue
            fy[k] = {"fy": k}; D["fy"].append(fy[k])
        if fy_start(k) >= fy_start(maxfy)-1:        # refresh latest two FY rows only
            agg = fy_by[k]
            fy[k].update(fy=k, eq=agg["eq"], debt=agg["tot"]-agg["eq"], tot=agg["tot"])
    D["fy"].sort(key=lambda r: fy_start(r["fy"]))
    # recompute cum ONLY for refreshed/new tail rows; anchor on prior row's
    # existing (published) cum so historical rounding is preserved untouched.
    refreshed = {k for k in fy_by if fy_start(k) >= fy_start(maxfy)-1}
    idxs = [i for i,r in enumerate(D["fy"]) if r["fy"] in refreshed]
    start = min(idxs) if idxs else len(D["fy"])
    c = D["fy"][start-1]["cum"] if start > 0 else 0
    for r in D["fy"][start:]:
        c += r["tot"]; r["cum"] = c

# ---------------------------------------------------------------- validate
def validate(D):
    yms = [r["ym"] for r in D["monthly"]]
    if yms != sorted(set(yms)):
        raise SystemExit("FATAL: monthly months not unique/sorted. Nothing written.")
    now = datetime.date.today(); cur = f"{now.year:04d}-{now.month:02d}"
    for r in D["monthly"]:
        if r["ym"] > cur: raise SystemExit(f"FATAL: future month {r['ym']}")
        if abs(r["eq"])>CEIL_MONTH or abs(r["tot"])>CEIL_MONTH:
            raise SystemExit(f"FATAL: month {r['ym']} out of range")
    for r in D["cy"]:
        if abs(r["tot"])>CEIL_YEAR: raise SystemExit(f"FATAL: CY {r['y']} out of range")
    for r in D["fy"]:
        if abs(r["tot"])>CEIL_YEAR: raise SystemExit(f"FATAL: FY {r['fy']} out of range")
    # cumulative chain self-consistent for the TAIL we maintain (recent years are
    # exact; older published rows carry harmless sub-crore rounding we never touch).
    fyl=D["fy"]
    for i in range(max(1,len(fyl)-5), len(fyl)):
        if abs((fyl[i-1]["cum"]+fyl[i]["tot"])-fyl[i]["cum"])>2:
            raise SystemExit(f"FATAL: cum chain broken at FY {fyl[i]['fy']}")

# ---------------------------------------------------------------- write
def write_all(html, span, D):
    blob = json.dumps(D, separators=(",",":"), ensure_ascii=False)
    new_html = html[:span[0]] + blob + html[span[1]:]
    open(INDEX, "w", encoding="utf-8").write(new_html)
    json.dump(D, open(JSON_MIRROR, "w", encoding="utf-8"), separators=(",",":"), ensure_ascii=False)
    try:
        write_xlsx(D)
    except Exception as e:
        log("  (xlsx regen skipped:", e, ")")

def write_xlsx(D):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Financial Year"
    ws.append(["Financial Year","Equity (Rs Cr)","Debt (Rs Cr)","Total (Rs Cr)","Cumulative (Rs Cr)"])
    for r in D["fy"]: ws.append([r["fy"], r["eq"], r.get("debt"), r["tot"], r["cum"]])
    ws2 = wb.create_sheet("Calendar Year")
    ws2.append(["Year","Equity (Rs Cr)","Debt+Other (Rs Cr)","Total (Rs Cr)","Months"])
    for r in D["cy"]: ws2.append([r["y"], r["eq"], r.get("dt"), r["tot"], r.get("m")])
    ws3 = wb.create_sheet("Monthly")
    ws3.append(["Month","Equity (Rs Cr)","Total (Rs Cr)"])
    for r in D["monthly"]: ws3.append([r["ym"], r["eq"], r["tot"]])
    wb.save(XLSX_OUT)

# ---------------------------------------------------------------- main
def main():
    html, span, D = read_D()
    before = (len(D["monthly"]), max(r["ym"] for r in D["monthly"]))
    csv_arg = None
    if "--from-csv" in sys.argv:
        csv_arg = sys.argv[sys.argv.index("--from-csv")+1]
    fresh = fetch_csv(csv_arg) if csv_arg else fetch_nsdl_months()
    log(f"Source returned {len(fresh)} month-rows "
        f"({min(fresh)}..{max(fresh)})" if fresh else "Source returned 0 rows")
    changed = apply_updates(D, fresh)
    validate(D)
    after = (len(D["monthly"]), max(r["ym"] for r in D["monthly"]))
    if not changed:
        log(f"No new/changed months. Latest still {before[1]}. Nothing to commit.")
        return 0
    log(f"Updated months: {changed}")
    log(f"  monthly: {before[0]} -> {after[0]} rows, latest {before[1]} -> {after[1]}")
    log(f"  current CY: {D['cy'][-1]}")
    log(f"  current FY: {D['fy'][-1]}")
    if DRY:
        log("DRY RUN — nothing written."); return 0
    write_all(html, span, D)
    log("Wrote index.html, fii-flows-data.json, FII_FPI_Historical_Data_India.xlsx")
    return 0

if __name__ == "__main__":
    sys.exit(main())
